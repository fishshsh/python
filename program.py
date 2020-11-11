from random import choice, randint 
from openpyxl import Workbook

camps = ['Огонь', 'Вода', 'Земля', 'Воздух']

names = []

with open('input.txt') as file:
    for line in file.readlines():
        names.append(line.strip())

wb = Workbook()

ws1 = wb.create_sheet("Взносы")
ws2 = wb.create_sheet("Поездки")

ws1.cell(row=1, column=1, value='ФИО')
ws1.cell(row=1, column=2, value='Кол-во взносов')
ws1.cell(row=1, column=3, value='Итоговая сумма')
ws1.cell(row=1, column=4, value='Решение комиссии')

ws2.cell(row=1, column=1, value='ФИО')
ws2.cell(row=1, column=2, value='Статус')
ws2.cell(row=1, column=3, value='Санаторий')

for i, name in enumerate(names, 2):
    payments_count = randint(6, 12)

    total_payments_sum = 0
    for j in range(payments_count):
        total_payments_sum += randint(100, 500)

    if payments_count == 12 and total_payments_sum > 3000:
        comission_decision = '+'
        status = 'Приглашен(-а)'
        camp = choice(camps)
    else:
        comission_decision = '-'
        status = 'Не приглашен(-а)'
        camp = '-'

    ws1.cell(row=i, column=1, value=name)
    ws1.cell(row=i, column=2, value=payments_count)
    ws1.cell(row=i, column=3, value=total_payments_sum)
    ws1.cell(row=i, column=4, value=comission_decision)

    ws2.cell(row=i, column=1, value=name)
    ws2.cell(row=i, column=2, value=status)
    ws2.cell(row=i, column=3, value=camp)

wb.save('output.xlsx')
